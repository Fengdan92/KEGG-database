import requests
from request_api_functions import *
import multi_pairs_functions as mpf
import datetime
import os

from openpyxl import load_workbook
import construct_network_functions as cnf


database_direct = 'your-KEGG-database-directory' # e.g. KEGG/database/

def info(title):
    print(title)
    print('module name:', __name__)
    print('parent process:', os.getppid())
    print('process id:', os.getpid())

    return None


def request_api():
	info('request api')

	r = requests.get('http://rest.kegg.jp/list/organism')
	txt_org = r.text.encode('ascii','ignore')

	lst_org = []
	for line in txt_org.splitlines():
		lst_org.append(dict(zip(('Code','Abbr','Name','Type'),line.split('\t')))) 

	filename_org = 'KEGG_organism_' + str(datetime.date.today()) + '.xlsx'
	savexls_org(dir=database_direct, filename=filename_org, lst_org=lst_org)

	lst_bact = []
	for org in lst_org:
		if 'Bacteria' in org['Type']:
			lst_bact.append(org)

	filename_bact = 'KEGG_bacteria_' + str(datetime.date.today()) + '.xlsx'
	savexls_bact(dir=database_direct, filename=filename_bact, lst_bact=lst_bact)

	print(">>> Creating list of abbrieviations...")
	abbr_of_bacteria = []
	for bact in lst_bact:
		abbr_of_bacteria.append(bact['Abbr'].upper())

	print('>>> Requesting enzyme list...')
	r = requests.get('http://rest.kegg.jp/list/enzyme')
	txt_enzyme = r.text.encode('ascii','ignore') 

	lst_enzyme = []
	for line in txt_enzyme.splitlines():
		lst_enzyme.append(dict(zip(('Code','Name'),line.split('\t'))))
		# this list may include deleted/transferred enzyme entries.

	print('>>> Creating list of EC numbers...')
	lst_ec = []
	for enzyme in lst_enzyme:
		lst_ec.append(enzyme['Code'].replace("ec:", ""))
		# this list may include deleted/transferred enzyme entries.

	'''For a specific EC number (e.g. 1.1.1.1), we use keyword "GENES" and "DBLINKS" to get the "Genes" entry of this enzyme, 
which contains information of abbrieviation of bacteria that encode this enzyme. We assume that
enzymes without GENES or DBLINKS entries are deleted ec entries. '''
	print('>>> Establishing EC - bacteria relations...')
	lst_ec_bact = [] 
	# a list of dictionaries that record EC number and which bacteria encode this enzyme

	i = 0
	for ec in lst_ec:
		address='http://rest.kegg.jp/get/'+ec
		r = requests.get(address)
		if 'GENES' in r.text:
			tmp_text1=r.text.split("GENES")[1]
			if 'DBLINKS' in tmp_text1:
				tmp_text2=tmp_text1.split("DBLINKS")[0] 
				lst_ec_bact.append({"EC":ec})
				for bact in abbr_of_bacteria:
					if bact in tmp_text2:
						lst_ec_bact[i][bact]='1'
					else:
						lst_ec_bact[i][bact]='0'
				i = i + 1

	filename_ec_bact = 'KEGG_EC_bact_' + str(datetime.date.today()) + '.xlsx'
	savexls_ec_bact(dir=database_direct, filename=filename_ec_bact, lst_ec_bact=lst_ec_bact, lst_abbr=abbr_of_bacteria)

	'''For a specific EC number (e.g. 1.1.1.1), we use keyword "ALL_REAC" and "SUBSTRATE" to get 
the "ALL_REAC" entry of this enzyme, which contains information of reactions this enzyme
catalyzes. We assume that enzymes without ALL_REAC or SUBSTRATE entries are deleted ec entries. '''
	print('>>> Establishing EC - reaction relations...')
	lst_ec_reac = [] 
	# a list of dictionaries that record EC number and what reactions this enzyme catalyzes.

	for ec in lst_ec:
		address='http://rest.kegg.jp/get/'+ec
		r = requests.get(address)
		if 'ALL_REAC' in r.text:
			tmp_text1=r.text.split('ALL_REAC')[1]
			if 'SUBSTRATE' in tmp_text1:
				tmp_text2=tmp_text1.split('SUBSTRATE')[0] 
				tmp_text2 = tmp_text2.encode('ascii','ignore')
				lst_reac=tmp_text2.split()
				lst_reac=[x for x in lst_reac if x.startswith('R')]
				lst_reac=[x.replace(';', '') for x in lst_reac]
				lst_ec_reac.append({"EC":ec, "Reactions":lst_reac})

	filename_ec_reac = 'KEGG_EC_reac_' + str(datetime.date.today()) + '.xlsx'
	savexls_ec_reac(dir=database_direct, filename=filename_ec_reac, lst_ec_reac=lst_ec_reac)

	return None


def obtain_dict(filename_ec_bact, filename_ec_reac, filename_db):
	
	print('>>> Reading EC-bacteria relations:\n')
	wb_ec_bact = load_workbook(filename = database_direct + filename_ec_bact)
	ws_ec_bact = wb_ec_bact['Sheet 1']
	lst_row1 = ws_ec_bact[1]
	lst_bact = [(x.value).encode('ascii','ignore') for x in lst_row1]
	lst_bact = lst_bact[1:]

	lst_ec_tmp = ws_ec_bact['A']
	lst_ec = [(x.value).encode('ascii','ignore') for x in lst_ec_tmp]
	lst_ec = lst_ec[1:]

	dict_bact_ec = {}
	for ba_col in ws_ec_bact.iter_cols(min_col = 2):
		ba_lst = [(x.value).encode('ascii','ignore') for x in ba_col]
		ba_name = ba_lst[0]
		print 'The bacterium being studied is: ' + ba_name
		ba_lst = ba_lst[1:]
		ba_ec_lst = []
		idx = 0
		for flag in ba_lst:
			if flag == '1':
				ba_ec_lst.append(lst_ec[idx])
			idx = idx + 1
		dict_bact_ec[ba_name]=ba_ec_lst


	wb_ec_reac = load_workbook(filename = database_direct + filename_ec_reac)
	ws_ec_reac = wb_ec_reac['Sheet 1']

	wb_db = load_workbook(filename = database_direct + filename_db)
	ws_db = wb_db['Reactions']
	ws_sp = wb_db['Non-enzymatic reactions']

	reac_column_a = ws_db['A']
	lst_reac = []
	for item in reac_column_a:
		if item.value == None:
			lst_reac.append('X')
		else:
			lst_reac.append((item.value).encode('ascii','ignore'))

	reac_col_sp = ws_sp['A']
	lst_reac_sp = [(x.value).encode('ascii','ignore') for x in reac_col_sp if x.value]
	lst_reac_sp = lst_reac_sp[2:]

	dict_ec_pairs = {}
	for row in ws_ec_reac.iter_rows(min_row = 2):
		row_values=[(x.value).encode('ascii','ignore') for x in row if x.value]
		ec_name = row_values[0]
		print 'The enzyme being studied is: ' + ec_name
		ec_reac_lst = row_values[1:]
		pairs_lst = mpf.obtain_pairs(ba_reac_lst=ec_reac_lst, reac_lst=lst_reac, ws = ws_db)
		pairs_lst_full = mpf.append_pairs_sp(reac_lst=lst_reac_sp, ws = ws_sp, pairs_lst = pairs_lst)
		dict_ec_pairs[ec_name]=pairs_lst_full

	return [lst_bact, dict_bact_ec, dict_ec_pairs]

