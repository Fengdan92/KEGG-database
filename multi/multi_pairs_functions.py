from openpyxl import load_workbook

def obtain_pairs(ba_reac_lst, reac_lst, ws):
	'''
	Obtain list of reactant pairs from enzymatic reactions for a certain bacterium.

	Parameters
	----------
	ba_reac_lst : list of strings
		list of reactions that this bacterium encodes
	reac_lst : list of strings
		List of reactions from the bioreaction database 'Reactions' tab column 'A'
	ws : openpyxl.worksheet.worksheet.Worksheet
		The worksheet that contains enzymatic reaction data.
		
	Returns
	-------
	pairs_lst : list of tuples - [(int, int), (int, int),...]
		List of reactant pairs from enzymatic reactions for a certain
		bacterium. The edges are *directed*.
	'''
	pairs_lst=[]
	for reac in ba_reac_lst:
		if reac in reac_lst:
			idx = reac_lst.index(reac)
			# below: index +1 due to different indexing style (python - 0 indexing, ws - 1 indexing)
			irreversibility = ws.cell(row=idx+1, column = 3).value
			irreversibility = int(irreversibility)
			if ws.cell(row=idx+1, column = 6).value:
				src = ws.cell(row=idx+1, column = 6).value
				dest = ws.cell(row=idx+1, column = 7).value
				if ((int(src), int(dest))) not in pairs_lst:
					pairs_lst.append((int(src), int(dest)))
				if irreversibility == 0:
				# if reaction reversible
					if ((int(dest), int(src))) not in pairs_lst:
						pairs_lst.append((int(dest), int(src)))
			if ws.cell(row=idx+1, column = 8).value:
				src = ws.cell(row=idx+1, column = 8).value
				dest = ws.cell(row=idx+1, column = 9).value
				if ((int(src), int(dest))) not in pairs_lst:
					pairs_lst.append((int(src), int(dest)))
				if irreversibility == 0:
				# if reaction reversible
					if ((int(dest), int(src))) not in pairs_lst:
						pairs_lst.append((int(dest), int(src)))
			if ws.cell(row=idx+1, column = 10).value:
				src = ws.cell(row=idx+1, column = 10).value
				dest = ws.cell(row=idx+1, column = 11).value
				if ((int(src), int(dest))) not in pairs_lst:
					pairs_lst.append((int(src), int(dest)))
				if irreversibility == 0:
				# if reaction reversible
					if ((int(dest), int(src))) not in pairs_lst:
						pairs_lst.append((int(dest), int(src)))
			if ws.cell(row=idx+1, column = 12).value:
				src = ws.cell(row=idx+1, column = 12).value
				dest = ws.cell(row=idx+1, column = 13).value
				if ((int(src), int(dest))) not in pairs_lst:
					pairs_lst.append((int(src), int(dest)))
				if irreversibility == 0:
				# if reaction reversible
					if ((int(dest), int(src))) not in pairs_lst:
						pairs_lst.append((int(dest), int(src)))
	
	return pairs_lst


def append_pairs_sp(reac_lst, ws, pairs_lst):
	'''
	Append list of reactant pairs (directed) from non-enzymatic reactions 
	Principle: A spontaneous reaction is included during reconstruction of an
	organism-specific metabolic network, only if one of its reactants is involved 
	in the enzymatic reactions known for this organism. See Stelzer et al. 2011.

	Parameters
	----------
	reac_lst : list of strings
		List of reactions from the bioreaction database 'non-enzymatic reactions' tab column 'A'
	ws : openpyxl.worksheet.worksheet.Worksheet
		The worksheet that contains non-enzymatic reaction data.
	pairs_lst : list of tuples - [(int, int), (int, int),...]
		List of reactant pairs from enzymatic reactions for a certain
		bacterium. The edges are *directed*.
		
	Returns
	-------
	pairs_lst : list of tuples - [(int, int), (int, int),...]
		The updated list of reactant pairs (directed) from both non-enzymatic reactions
		and enzymatic reactions for this bacterium.
	'''
	lst_nodes = [x for y in pairs_lst for x in y]
	lst_nodes = list(set(lst_nodes))

	for reac in reac_lst:
		idx = reac_lst.index(reac)
		# +3 due to different indexing style (python - 0 indexing, ws - 1 indexing)
		# and because there are two rows of headers
		pairs_lst_tmp = []
		# contains all pairs of reactant for this reaction
		irreversibility = ws.cell(row=idx+3, column = 3).value
		irreversibility = int(irreversibility)
		if ws.cell(row=idx+3, column = 6).value:
			src = ws.cell(row=idx+3, column = 6).value
			dest = ws.cell(row=idx+3, column = 7).value
			pairs_lst_tmp.append((int(src), int(dest)))
			if irreversibility == 0:
			# if reaction reversible
				pairs_lst_tmp.append((int(dest), int(src)))
		if ws.cell(row=idx+3, column = 8).value:
			src = ws.cell(row=idx+3, column = 8).value
			dest = ws.cell(row=idx+3, column = 9).value
			pairs_lst_tmp.append((int(src), int(dest)))
			if irreversibility == 0:
			# if reaction reversible
				pairs_lst_tmp.append((int(dest), int(src)))
		if ws.cell(row=idx+3, column = 10).value:
			src = ws.cell(row=idx+3, column = 10).value
			dest = ws.cell(row=idx+3, column = 11).value
			pairs_lst_tmp.append((int(src), int(dest)))
			if irreversibility == 0:
			# if reaction reversible
				pairs_lst_tmp.append((int(dest), int(src)))
		if ws.cell(row=idx+3, column = 12).value:
			src = ws.cell(row=idx+3, column = 12).value
			dest = ws.cell(row=idx+3, column = 13).value
			pairs_lst_tmp.append((int(src), int(dest)))
			if irreversibility == 0:
			# if reaction reversible
				pairs_lst_tmp.append((int(dest), int(src)))

		lst_nodes_tmp = [x for y in pairs_lst_tmp for x in y]
		lst_nodes_tmp = list(set(lst_nodes_tmp))

		for node in lst_nodes_tmp:
			if node in lst_nodes:
				for tp in pairs_lst_tmp:
					if tp not in pairs_lst:
						pairs_lst.append(tp)
	
	return pairs_lst