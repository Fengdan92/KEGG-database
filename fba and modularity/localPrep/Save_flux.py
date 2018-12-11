from openpyxl import load_workbook
import pickle

wb = load_workbook(filename = 'Flux_GluAerobic.xlsx')
ws = wb['Sheet1']

# list of reaction abbreviations
rxn_abbr = ws['A']
rxn_abbr = [x.value for x in rxn_abbr]
rxn_abbr = rxn_abbr[1:]

# fluxes
rxn_flux = ws['C']
rxn_flux = rxn_flux[1:]
rxn_flux = [float(x.value) for x in rxn_flux]

dict_flux = {}
for i in range(len(rxn_abbr)):
	# eliminate reactions that have flux smaller than 1e-5
	# also eliminate the BIOMASS reactions, as the unit of flux through it is h-1.
	if (abs(rxn_flux[i]) >= 1e-5) and ('BIOMASS' not in rxn_abbr[i]):
		dict_flux[rxn_abbr[i]]=rxn_flux[i]

# save as pickle
file = open(r'dict_flux.pkl', 'wb')
pickle.dump(dict_flux, file)
file.close()