from openpyxl import load_workbook
from xlsxwriter import Workbook
import pickle

wb = load_workbook(filename='iML1515.xlsx')
ws = wb['Reaction List']

# list of reaction abbreviations
rxn_abbr = ws['A']
rxn_abbr = [x.value for x in rxn_abbr]
rxn_abbr = rxn_abbr[1:]

# list of reaction equations
rxn_equ = ws['C']
rxn_equ = [x.value for x in rxn_equ]
rxn_equ = rxn_equ[1:]

# obtain pairs
dict_all = {}
for i in range(len(rxn_abbr)):
	dict_this_reaction = {}
	abbr = rxn_abbr[i]
	equ = rxn_equ[i]
	if '->' in equ:
		twoSides = equ.split('->')
	if '<=>' in equ:
		twoSides = equ.split('<=>')
	sources = twoSides[0]
	dest = twoSides[1]
	lst_src = sources.split('+')
	lst_dest = dest.split('+')
	lst_src = [ x.strip() for x in lst_src ]
	lst_dest = [ x.strip() for x in lst_dest ]
	
	for srcnode in lst_src:
		if ' ' in srcnode:
			lst = srcnode.split(' ')
			weight = float(lst[0])
			node = lst[1]
		else:
			node = srcnode
			weight = 1.0
		dict_this_reaction[(node,abbr)] = weight
	
	for destnode in lst_dest:
		if ' ' in destnode:
			lst = destnode.split(' ')
			weight = float(lst[0])
			node = lst[1]
		else:
			node = destnode
			weight = 1.0
		if node != '':
			dict_this_reaction[(abbr, node)] = weight

	dict_all[abbr]=dict_this_reaction

# save as pickle
file = open(r'dict_all.pkl', 'wb')
pickle.dump(dict_all, file)
file.close()

# save to xlsx
header_lst = ['Abbr','Equ','Pairs']
wb_write = Workbook('iML1515_pairs.xlsx')
ws_write = wb_write.add_worksheet('Reactions')

header_row = 0
for col in range(0,len(header_lst)):
	ws_write.write(header_row, col, header_lst[col])

row = 1
for i in range(len(rxn_abbr)):
	abbr = rxn_abbr[i]
	equ = rxn_equ[i]
	ws_write.write(row, 0, abbr)
	ws_write.write(row, 1, equ)
	dict_this_reaction = dict_all[abbr]
	string_pairs = ', '.join("{!s}={!r}".format(key,val) for (key,val) in dict_this_reaction.items())
	ws_write.write(row, 2, string_pairs)
	row += 1

wb_write.close()